#!/usr/bin/env python3
"""
Windows-Safe Document Extraction System
Fixes PermissionError with temp files
"""

import zipfile
import os
import shutil
import json
import fitz  # PyMuPDF
import camelot
import pandas as pd
from paddleocr import PaddleOCR
from collections import defaultdict
from pathlib import Path
import docx
import logging
import tempfile
import time
from typing import List, Dict, Any
import openpyxl

# ---------- WINDOWS-SAFE LOGGING ----------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ---------- OCR ENGINE ----------
try:
    ocr_engine = PaddleOCR(use_textline_orientation=True, lang="en")
    logger.info("✅ PaddleOCR initialized")
except Exception as e:
    logger.warning(f"⚠️ PaddleOCR init failed: {e}")
    ocr_engine = None

# ---------- SUPPORTED FORMATS ----------
SUPPORTED_FORMATS = {'.pdf', '.docx', '.doc', '.xlsx', '.xls', '.png', '.jpg', '.jpeg', '.bmp', '.tiff'}

def safe_rmtree(path, retries=5, delay=0.5):
    """Windows-safe directory deletion with retries"""
    for attempt in range(retries):
        try:
            if os.path.exists(path):
                shutil.rmtree(path)
            return True
        except Exception as e:
            logger.warning(f"Delete attempt {attempt+1} failed: {e}")
            if attempt < retries - 1:
                time.sleep(delay)
    logger.error(f"❌ Failed to delete {path} after {retries} attempts")
    return False

def extract_zip(zip_path, out_dir="extracted_files"):
    """Extract zip with safe cleanup"""
    safe_rmtree(out_dir)
    os.makedirs(out_dir)
    
    try:
        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(out_dir)
        logger.info(f"✅ Extracted zip: {zip_path}")
    except Exception as e:
        logger.error(f"❌ Failed to extract zip: {e}")
        return []
    
    # Find supported files
    all_files = []
    for root, dirs, files in os.walk(out_dir):
        for file in files:
            if Path(file).suffix.lower() in SUPPORTED_FORMATS:
                all_files.append(os.path.join(root, file))
    
    return all_files

def is_digital_pdf(pdf_path):
    """Check if PDF is digital"""
    try:
        doc = fitz.open(pdf_path)
        text_blocks = 0
        total_blocks = 0
        
        for page in doc:
            blocks = page.get_text("blocks")
            for block in blocks:
                if block[6] == 0:
                    total_blocks += 1
                    if block[4].strip():
                        text_blocks += 1
        
        doc.close()
        return text_blocks > (total_blocks * 0.3)
    except Exception:
        return False

def safe_cleanup(temp_files):
    """Safely delete temp files with retries"""
    for file_path in temp_files:
        for attempt in range(3):
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                break
            except Exception as e:
                if attempt == 2:
                    logger.warning(f"Failed to delete {file_path}: {e}")
                time.sleep(0.1)

def extract_tables_with_fallback(file_path, file_type, page_num=None):
    """Extract tables"""
    tables = []
    
    if file_type == "pdf":
        try:
            if page_num:
                camelot_tables = camelot.read_pdf(file_path, pages=str(page_num), flavor="lattice")
            else:
                camelot_tables = camelot.read_pdf(file_path, pages="all", flavor="lattice")
            
            for t in camelot_tables:
                df = t.df.fillna("")
                # Convert to strings to preserve formatting
                table_data = [[str(cell) if cell != "" else "" for cell in row] for row in df.values.tolist()]
                if table_data and len(table_data) > 1:
                    tables.append(table_data)
        except:
            pass
        
        if not tables:
            try:
                if page_num:
                    camelot_tables = camelot.read_pdf(file_path, pages=str(page_num), flavor="stream")
                else:
                    camelot_tables = camelot.read_pdf(file_path, pages="all", flavor="stream")
                
                for t in camelot_tables:
                    df = t.df.fillna("")
                    # Convert to strings to preserve formatting
                    table_data = [[str(cell) if cell != "" else "" for cell in row] for row in df.values.tolist()]
                    if table_data and len(table_data) > 1:
                        tables.append(table_data)
            except:
                pass
    
    elif file_type == "excel":
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            # Convert to strings to preserve formatting
            header = [str(col) for col in df.columns.tolist()]
            table_data = [[str(cell) if not pd.isna(cell) else "" for cell in row] for row in df.values.tolist()]
            if table_data:
                tables.append([header] + table_data)
        except:
            pass
    
    return tables

def extract_digital_pdf(pdf_path):
    """Extract digital PDF"""
    try:
        doc = fitz.open(pdf_path)
        pages_data = []
        
        tables_by_page = defaultdict(list)
        try:
            camelot_tables = camelot.read_pdf(pdf_path, pages="all", flavor="lattice")
            for t in camelot_tables:
                df = t.df.fillna("")
                # Convert to strings to preserve formatting
                table_data = [[str(cell) if cell != "" else "" for cell in row] for row in df.values.tolist()]
                tables_by_page[t.page].append(table_data)
        except:
            try:
                camelot_tables = camelot.read_pdf(pdf_path, pages="all", flavor="stream")
                for t in camelot_tables:
                    df = t.df.fillna("")
                    # Convert to strings to preserve formatting
                    table_data = [[str(cell) if cell != "" else "" for cell in row] for row in df.values.tolist()]
                    tables_by_page[t.page].append(table_data)
            except:
                pass
        
        for i, page in enumerate(doc, start=1):
            blocks = page.get_text("blocks")
            blocks = sorted(blocks, key=lambda b: (b[1], b[0]))
            
            text_parts = []
            for block in blocks:
                if block[6] == 0:
                    text = block[4].strip()
                    if len(text) > 2:
                        text_parts.append(text)
            
            pages_data.append({
                "page_no": i,
                "text": "\n\n".join(text_parts),
                "tables": tables_by_page.get(i, [])
            })
        
        doc.close()
        return pages_data
    except Exception as e:
        logger.error(f"❌ Digital PDF failed: {e}")
        return []

def extract_scanned_pdf(pdf_path):
    """Extract scanned PDF"""
    try:
        doc = fitz.open(pdf_path)
        pages_data = []
        temp_files = []
        
        for i, page in enumerate(doc, start=1):
            pix = page.get_pixmap(dpi=300)
            img_path = f"temp_page_{i}_{int(time.time())}.png"
            pix.save(img_path)
            temp_files.append(img_path)
            
            try:
                if ocr_engine:
                    ocr_result = ocr_engine.ocr(img_path, cls=True)
                    lines = [line[1][0] for line in ocr_result[0]] if ocr_result and ocr_result[0] else []
                    text = "\n".join(lines)
                else:
                    text = "OCR unavailable"
            except:
                text = ""
            
            pages_data.append({
                "page_no": i,
                "text": text,
                "tables": []
            })
        
        doc.close()
        safe_cleanup(temp_files)
        return pages_data
    except Exception as e:
        logger.error(f"❌ Scanned PDF failed: {e}")
        return []

def extract_docx(docx_path):
    """Extract DOCX"""
    try:
        doc = docx.Document(docx_path)
        text_parts = []
        tables_data = []
        
        for element in doc.element.body:
            if element.tag.endswith('p'):
                para = docx.oxml.parse_xml(docx.oxml.serialize_for_reading(element))
                text = "".join(node.text for node in para.iter() if node.text)
                if text.strip():
                    text_parts.append(text.strip())
            
            elif element.tag.endswith('tbl'):
                table_rows = []
                for row in element.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tr'):
                    cells = [''.join(node.text for node in cell.iter() if node.text).strip() 
                            for cell in row.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tc')]
                    if cells:
                        table_rows.append(cells)
                if table_rows:
                    tables_data.append(table_rows)
        
        return [{
            "page_no": 1,
            "text": "\n\n".join(text_parts),
            "tables": tables_data
        }]
    except Exception as e:
        logger.error(f"❌ DOCX failed: {e}")
        return []

def extract_excel(excel_path):
    """Extract Excel"""
    try:
        if isinstance(excel_path, str):
            wb = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)
        else:
             # Handle file-like objects if necessary, though current usage ignores this
             wb = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)
             
        all_sheets = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = []
            
            # Iterate through rows
            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    val = cell.value
                    if val is None:
                        row_data.append("")
                        continue

                    # Handle different value types
                    if isinstance(val, str):
                        # String values - preserve as-is
                        row_data.append(val)
                    elif isinstance(val, (int, float)):
                        # Numeric values - check formatting
                        fmt = cell.number_format
                        formatted_val = str(val)
                        
                        # Check if format is not General and has relevant chars
                        if fmt and fmt.lower() != 'general':
                            import re
                            
                            # Detect decimal precision from format string
                            precision_match = re.search(r'\.([0#]+)', fmt)
                            
                            # Extract text literals (e.g., "Cr", "Dr") from format
                            # Format strings can have quoted text like:
                            # - 0.00 "Cr"
                            # - ""0.00" Cr" (with empty quotes)
                            # Find all quoted strings, filter out empty ones
                            all_quotes = re.findall(r'"([^"]*)"', fmt)
                            text_literals = [t for t in all_quotes if t]  # Remove empty strings
                            
                            if precision_match:
                                precision = len(precision_match.group(1))
                                # Format with fixed precision
                                formatted_val = f"{val:.{precision}f}"
                            
                            # Append text literals if found
                            if text_literals:
                                # Join all text literals with space
                                formatted_val = formatted_val + " " + " ".join(text_literals)
                        
                        row_data.append(formatted_val)
                    else:
                        # Handle other types (dates, booleans, etc.)
                        row_data.append(str(val))
                
                # Check if row has any content
                if any(row_data):
                    sheet_data.append(row_data)

            if sheet_data:
                all_sheets.append(sheet_data)
        
        wb.close()
        
        return [{
            "page_no": 1,
            "text": f"Excel file with {len(all_sheets)} sheets",
            "tables": all_sheets
        }]
    except Exception as e:
        logger.error(f"❌ Excel failed: {e}")
        return []

def extract_image(image_path):
    """Extract image"""
    try:
        if not ocr_engine:
            return [{"page_no": 1, "text": "OCR unavailable", "tables": []}]
        
        ocr_result = ocr_engine.ocr(image_path, cls=True)
        lines = [line[1][0] for line in ocr_result[0]] if ocr_result and ocr_result[0] else []
        
        return [{
            "page_no": 1,
            "text": "\n".join(lines),
            "tables": []
        }]
    except Exception as e:
        logger.error(f"❌ Image failed: {e}")
        return []

def get_file_type(file_path):
    """Get file type"""
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf": return "pdf"
    elif ext in [".docx", ".doc"]: return "docx"
    elif ext in [".xlsx", ".xls"]: return "excel"
    elif ext in [".png", ".jpg", ".jpeg", ".bmp", ".tiff"]: return "image"
    return "unknown"

def extract_document(file_path):
    """Main extraction function"""
    file_type = get_file_type(file_path)
    file_name = os.path.basename(file_path)
    
    if file_type == "pdf":
        pdf_type = "digital" if is_digital_pdf(file_path) else "scanned"
        pages = extract_digital_pdf(file_path) if pdf_type == "digital" else extract_scanned_pdf(file_path)
        return {
            "file_name": file_name,
            "pdf_type": pdf_type,
            "total_pages": len(pages),
            "pages": pages
        }
    
    elif file_type == "docx":
        pages = extract_docx(file_path)
        return {"file_name": file_name, "pdf_type": "docx", "total_pages": 1, "pages": pages}
    
    elif file_type == "excel":
        pages = extract_excel(file_path)
        return {"file_name": file_name, "pdf_type": "excel", "total_pages": 1, "pages": pages}
    
    elif file_type == "image":
        pages = extract_image(file_path)
        return {"file_name": file_name, "pdf_type": "image", "total_pages": 1, "pages": pages}
    
    return None

def process_zip(zip_path, output_file="extracted_data.json"):
    """Process zip file"""
    try:
        files = extract_zip(zip_path)
        if not files:
            logger.error("❌ No supported files")
            return
        
        logger.info(f"🔍 Found {len(files)} files")
        
        output = []
        for i, file_path in enumerate(files, 1):
            logger.info(f"[{i}/{len(files)}] {os.path.basename(file_path)}")
            result = extract_document(file_path)
            if result:
                output.append(result)
        
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        
        logger.info(f"\n✅ SUCCESS!")
        logger.info(f"📄 {output_file}")
        logger.info(f"📊 {len(output)} documents")
        return output
    
    except Exception as e:
        logger.error(f"❌ Failed: {e}")
        return None

def process_single_file(file_path, output_file="extracted_data.json"):
    """Process single file"""
    try:
        logger.info(f"Processing: {os.path.basename(file_path)}")
        result = extract_document(file_path)
        if result:
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump([result], f, indent=2, ensure_ascii=False)
            logger.info(f"✅ {output_file}")
            return [result]
        return None
    except Exception as e:
        logger.error(f"❌ Failed: {e}")
        return None

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        input_path = sys.argv[1]
        if input_path.lower().endswith(".zip"):
            process_zip(input_path)
        else:
            process_single_file(input_path)
    else:
        print("Usage: python tds_extractor_v2.py <file_or_zip>")
        print("Example: python tds_extractor_v2.py input.zip")
